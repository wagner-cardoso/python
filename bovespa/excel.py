# Criar um arquivo do Excel

from openpyxl import Workbook

arquivo_excel = Workbook()
planilha1 = arquivo_excel.active

planilha1.title = "Acoes"
planilha2 = arquivo_excel.create_sheet("Fiis")

print(arquivo_excel.sheetnames)

planilha1['A1'] = "DATA"
planilha1['B1'] = "CODIGO"
planilha1['C1'] = "OPERACAO"  # compra ou venda
planilha1['D1'] = "QTD"
planilha1['E1'] = "VALOR"
planilha1['F1'] = "TOTAL"

# arquivo_excel.save("bovespa.xlsx")